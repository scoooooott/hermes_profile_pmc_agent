#!/usr/bin/env python3
"""PMC 数据管线 — Excel → DuckDB ODS 导入
三类文件 → 三目录 → 7 张 ODS 表

用法:
    cd ~/workspace/pmc-agents
    python3 scripts/pmc_import.py

前提: API 服务在 localhost:8765 运行，Excel 文件已下载到 ~/pmc-data/
"""

import os, re
from datetime import datetime

PMC_DIR = os.path.expanduser("~/pmc-data")
DB_PATH = os.path.expanduser("os.path.expanduser(os.environ.get("PMC_DB_PATH", "~/pmc-data/pmc_ods.duckdb"))")


def sanitize(val):
    """清洗 Excel 不允许的控制字符"""
    if isinstance(val, str):
        return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', val)
    return val


def read_excel(path):
    import openpyxl
    return openpyxl.load_workbook(path, data_only=True)


def import_sheet(con, wb, sheet_name, table_name, mode, upsert_keys=None):
    """mode: 'overwrite' | 'upsert'"""
    if sheet_name not in wb.sheetnames:
        print(f"  ⚠️  Sheet '{sheet_name}' 不存在，跳过")
        return 0

    ws = wb[sheet_name]
    headers = [sanitize(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]

    rows = []
    for r in range(3, ws.max_row + 1):
        row = [sanitize(ws.cell(r, c).value) for c in range(1, len(headers) + 1)]
        if all(v is None or v == '' for v in row):
            continue
        rows.append(row)

    if not rows:
        print(f"  ⚠️  Sheet '{sheet_name}' 无数据，跳过")
        return 0

    col_map = {
        "SKU编码": "sku_code", "SPU编码": "spu_code", "品名": "product_name",
        "类目": "category", "品牌": "brand", "上架日期": "launch_date",
        "上架状态": "status", "生命周期": "lifecycle", "货盘等级": "tier",
        "生产周期天数": "production_cycle_days",
        "人工日均销目标值": "manual_daily_sale_target",
        "MOQ(最小起订量)": "moq", "Lead Time(采购交期,天)": "lead_time",
        "updated_at": "updated_at",
        "销售日期": "sale_date", "日销量": "daily_qty",
        "最小销售单元": "msu_id", "最小销售单元ID": "msu_id",
        "采购单号": "po_number", "下单日期": "order_date",
        "采购数量": "order_qty", "预计到货日期": "eta",
        "收货日期": "receipt_date", "收货数量": "receipt_qty",
        "收货仓库": "warehouse_id", "供给仓库ID": "warehouse_id",
        "物流单号": "tracking_number", "发货日期": "ship_date",
        "发货数量": "ship_qty", "收获仓库": "dest_warehouse",
        "可售库存": "inv_available", "锁定库存": "inv_locked",
        "在途库存": "inv_onway",
        "国内库存": "inv_domestic", "采购在途": "inv_purchase_onway",
        "采购在单": "inv_inorder", "快照时间": "snapshot_time",
        "店铺": "shop", "仓库": "warehouse_code",
        "参数编号": "param_no", "参数ID": "param_id", "参数名称": "param_name",
        "数据类型/取值范围": "param_type", "默认值": "param_default", "备注": "param_note",
    }

    cols = [col_map.get(h, f'col_{i}') for i, h in enumerate(headers)]
    cols = [re.sub(r'[^a-z0-9_]', '_', c.lower()) for c in cols]

    tmp = f"_tmp_{table_name}"
    placeholders = ', '.join(['?' for _ in cols])
    col_defs = ', '.join(f'"{c}" VARCHAR' for c in cols)

    if mode == 'overwrite':
        con.execute(f"DROP TABLE IF EXISTS {table_name}")
        con.execute(f"DROP TABLE IF EXISTS {tmp}")
        con.execute(f'CREATE TABLE "{tmp}" ({col_defs})')
        con.executemany(f'INSERT INTO "{tmp}" VALUES ({placeholders})', rows)
        con.execute(f'DROP TABLE IF EXISTS "{table_name}"')
        con.execute(f'ALTER TABLE "{tmp}" RENAME TO "{table_name}"')
    else:
        existing = con.execute(
            f"SELECT name FROM sqlite_master WHERE type='table' AND name='{table_name}'"
        ).fetchone()
        if not existing:
            con.execute(f'CREATE TABLE "{table_name}" ({col_defs})')
        if upsert_keys:
            key_cols = ', '.join(f'"{col_map.get(k, k)}"' for k in upsert_keys)
            try:
                con.executemany(
                    f'INSERT INTO "{table_name}" VALUES ({placeholders}) '
                    f'ON CONFLICT ({key_cols}) DO NOTHING', rows
                )
            except Exception:
                con.execute(
                    f'CREATE UNIQUE INDEX IF NOT EXISTS idx_{table_name}_upsert '
                    f'ON "{table_name}" ({key_cols})'
                )
                con.executemany(f'INSERT OR IGNORE INTO "{table_name}" VALUES ({placeholders})', rows)
        else:
            con.executemany(f'INSERT INTO "{table_name}" VALUES ({placeholders})', rows)

    return len(rows)


def main():
    import duckdb
    con = duckdb.connect(DB_PATH)
    print(f"PMC DuckDB import @ {datetime.now().isoformat()}")
    print(f"DB: {DB_PATH}\n")

    results = {}

    # static/ 全量覆盖
    for fname, sheet, table in [
        ("商品主数据.xlsx", "商品主数据", "ods_skus"),
        ("参数配置.xlsx", "参数配置", "ods_params"),
        ("SKU-最小销售单元-仓库映射.xlsx", "MSU映射", "ods_wmap"),
    ]:
        path = f"{PMC_DIR}/static/{fname}"
        if not os.path.exists(path):
            results[f"static/{fname}"] = "❌ 文件不存在"
            continue
        wb = read_excel(path)
        n = import_sheet(con, wb, sheet, table, "overwrite")
        results[f"static/{fname}"] = f"✅ {n:,} rows → {table}"

    # snapshot/ 全量覆盖（2 Sheet）
    snap = f"{PMC_DIR}/snapshot/库存快照.xlsx"
    if os.path.exists(snap):
        wb = read_excel(snap)
        n1 = import_sheet(con, wb, "国内库存快照", "ods_inventory_domestic", "overwrite")
        n2 = import_sheet(con, wb, "海外库存快照", "ods_inventory_overseas", "overwrite")
        results["snapshot/库存快照.xlsx"] = f"✅ 国内={n1:,} 海外={n2:,} rows"

    # incremental/ 增量追加（5 Sheet）
    inc = f"{PMC_DIR}/incremental/日增量数据.xlsx"
    if os.path.exists(inc):
        wb = read_excel(inc)
        for sheet, table, keys in [
            ("日销量", "ods_sales", ("销售日期", "SKU编码", "最小销售单元")),
            ("采购明细", "ods_po", ("下单日期", "采购单号", "SKU编码")),
            ("采购收货明细", "ods_po_recv", ("收货日期", "采购单号", "SKU编码")),
            ("海外发货明细", "ods_ship", ("发货日期", "物流单号", "SKU编码")),
            ("海外收货明细", "ods_ship_recv", ("收货日期", "物流单号", "SKU编码")),
        ]:
            n = import_sheet(con, wb, sheet, table, "upsert", keys)
            results[f"inc/{sheet}"] = f"✅ {n:,} rows → {table}"

    print("=" * 60)
    for k, v in results.items():
        print(f"  {k}: {v}")

    print("\n─ 各表行数 ─")
    for row in con.execute(
        "SELECT table_name FROM duckdb_tables() WHERE table_name LIKE 'ods_%' ORDER BY table_name"
    ).fetchall():
        t = row[0]
        cnt = con.execute(f'SELECT COUNT(*) FROM "{t}"').fetchone()[0]
        print(f"  {t}: {cnt:,} rows")

    con.close()
    print("\n✅ Import complete.")


if __name__ == "__main__":
    main()
