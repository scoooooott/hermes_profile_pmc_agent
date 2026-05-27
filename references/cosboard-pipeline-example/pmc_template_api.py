#!/usr/bin/env python3
"""PMC Excel 模板生成 API — cosboard 只读查询
启动: uvicorn pmc_template_api:app --host 0.0.0.0 --port 8765
"""

import io, os
import pymysql
from collections import defaultdict
from datetime import datetime
from typing import Optional
from fastapi import FastAPI, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = FastAPI(title="PMC Excel Template API", version="1.0")

COSBOARD = {
    "host": "8.134.131.227", "port": 3630,
    "user": "cosdbusr", "password": "CosBoard+Serv2025",
    "database": "cosboard", "charset": "utf8mb4"
}

EMPTY_NOTE = "【cosboard无此字段】"
MANUAL_NOTE = "【客户手动填写】"
CALC_NOTE = "【需系统计算派生】"

HEADER_FONT = Font(bold=True, size=11)
NOTE_FONT = Font(color="FF0000", size=9, italic=True)
EXPLAIN_FONT = Font(color="808080", size=9, italic=True)
HEADER_FILL = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
GAP_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def get_conn():
    conn = pymysql.connect(connect_timeout=10, read_timeout=600, **COSBOARD)
    with conn.cursor() as cur:
        cur.execute("SET SESSION net_read_timeout = 600")
        cur.execute("SET SESSION net_write_timeout = 600")
    return conn


def query(sql: str) -> list:
    """安全的只读查询——只允许 SELECT"""
    s = sql.strip().upper()
    if not s.startswith("SELECT") and not s.startswith("SHOW") and not s.startswith("DESCRIBE"):
        raise ValueError(f"禁止非查询操作: {sql[:50]}...")
    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute(sql)
        cols = [d[0] for d in cur.description]
        rows = cur.fetchall()
        return [dict(zip(cols, r)) for r in rows]
    finally:
        conn.close()



from urllib.parse import quote

def make_xlsx_response(wb: Workbook, filename: str):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    encoded = quote(filename)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"}
    )


def sanitize_cell(val):
    """清除 Excel 不允许的控制字符 (0x00-0x08, 0x0B-0x0C, 0x0E-0x1F)"""
    if isinstance(val, str):
        import re
        return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', val)
    return val


def write_sheet(ws, headers: list, explains: list, rows: list, gap_cols: set = None):
    """写标准 PMC 模板格式: Row1=表头, Row2=解释, Row3+=数据"""
    gap_cols = gap_cols or set()
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=sanitize_cell(h))
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        if h.endswith("】"):
            cell.font = Font(bold=True, size=11, color="FF0000")
            cell.fill = GAP_FILL

    for c, e in enumerate(explains, 1):
        cell = ws.cell(row=2, column=c, value=sanitize_cell(e))
        cell.font = EXPLAIN_FONT
        cell.alignment = Alignment(wrap_text=True)
        cell.border = border
        if c - 1 in gap_cols:
            cell.fill = GAP_FILL

    for r, row in enumerate(rows, 3):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=sanitize_cell(v))
            cell.border = border
            if c - 1 in gap_cols:
                cell.fill = GAP_FILL

    # auto width
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 20


# ============================================================
# Endpoints
# ============================================================

@app.get("/")
def root():
    return {"service": "PMC Excel Template API", "endpoints": [
        "/template/skus", "/template/params", "/template/msu-map",
        "/template/inventory", "/template/daily-data"
    ]}


@app.get("/template/skus")
def template_skus(brand: Optional[str] = Query(None, description="品牌过滤")):
    """商品主数据模板 — 14 列对齐 v3"""
    sql = """select d.sku, d.spu, d.name, d.category, d.listdate,
       case when date(d.listdate)>='2025-06-01' then '新品'
            when date(d.listdate)<'2024-01-01' then '衰退'
            else '成长' end as lifecycle_hint,
       e.tag_name as tier
from cdm_sku d
left join cdm_spu_tag e on d.spu = e.commodity_code
order by d.listdate desc"""
    data = query(sql)
    if brand:
        data = [r for r in data if brand.lower() in str(r.get('name','')).lower()]

    headers = ["SKU编码", "SPU编码", "品名", "类目", "品牌",
               "上架日期", "上架状态", "生命周期", "货盘等级",
               "生产周期天数", "人工日均销目标值",
               "MOQ(最小起订量)", "Lead Time(采购交期,天)", "updated_at"]
    explains = ["唯一标识", "款式编码", "商品名称", "商品类目", "商品品牌名称",
                "首次上架", "在售/下架", "新品/成长/成熟/衰退",
                "S=爆款/A=畅销/B=平销/C=长尾/N=新品,客户设初始值(场景00每日刷新)",
                "标准生产天数", "有值覆盖自动（留空=自动）",
                "供应商最小起订件数(未填=无限制)", "采购下单到入库标准天数(不含物流,未填=使用默认值30)",
                "更新时间"]
    gap_cols = {4, 6, 7, 9, 10, 11, 12}  # 品牌/上架状态/生命周期/生产周期/日均销/MOQ/LeadTime

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rows = [[
        r.get('sku',''), r.get('spu',''), r.get('name',''), r.get('category',''),
        '',  # 品牌 — cosboard 无此字段
        str(r.get('listdate',''))[:10], '',  # 上架状态需从其他表查
        r.get('lifecycle_hint',''), r.get('tier','B'),
        '', '', '', '',  # 生产周期/人工日均销/MOQ/LeadTime
        now
    ] for r in data]

    wb = Workbook()
    ws = wb.active; ws.title = "商品主数据"
    write_sheet(ws, headers, explains, rows, gap_cols)
    return make_xlsx_response(wb, "商品主数据.xlsx")


@app.get("/template/params")
def template_params():
    """参数配置模板"""
    headers = ["参数编号", "参数ID", "参数名称", "数据类型/取值范围", "默认值", "备注"]
    explains = ["编号", "英文标识", "中文名称", "类型或枚举", "默认值", "说明"]

    params = [
        ["P1", "stock_sale_ratio", "存销比目标", "按S/A/B/C/N浮点数", '{"S":1.5,"A":1.2,"B":1.0,"C":0.8,"N":1.0}', "安全库存覆盖月数"],
        ["P2", "segment_target_days", "库存段目标天数", "按S/A/B/C/N×4段JSON", '{"S":{"可售":15,"在途":20,"国内":30,"采购在单":45}}', "场景07库存结构分析消费"],
        ["P3", "forecast_params", "销售预测参数", "dict: 趋势阈值+系数+窗口+促销+季节性", '{"trend_window":30,"threshold":0.15,"boost_coef":1.2}', "场景01销量预测消费"],
        ["P4", "cycle_target_days", "周期目标值", "按货盘×{采购,海外发货}", '{"S":{"采购":15,"海外发货":25},"A":{"采购":20,"海外发货":30}}', "场景09周期分析（3周期）"],
        ["P5", "slow_moving_days", "呆滞天数目标值", "按S/A/B/C/N正整数", '{"S":60,"A":75,"B":90,"C":105,"N":120}', "场景02/06呆滞判定"],
        ["P6", "otb_limit", "OTB总量上限", "正整数", "50000", "场景08总量控制"],
        ["P7", "safety_stock_days", "安全库存天数", "按S/A/B/C/N整数(场景02膨胀基底)", '{"S":30,"A":25,"B":20,"C":15,"N":20}', "场景02库存缺口膨胀消费"],
        ["P8", "procurement_cycle_target", "备货周期目标", "按S/A/B/C/N整数(天数)", '{"S":7,"A":10,"B":14,"C":21,"N":10}', "场景04备货Lead Time参考"],
        ["P9", "replenish_trigger_threshold", "补货触发阈值", "按货盘整数(海外库存天数低于此触发补货)", '{"S":15,"A":12,"B":10,"C":8,"N":10}', "场景05补货触发消费"],
        ["P10", "target_overseas_inv_days", "目标海外库存天数", "按货盘整数", '{"S":30,"A":25,"B":20,"C":15,"N":20}', "场景05补货目标消费"],
        ["P11", "daily_sale_override_strategy", "日均销覆盖策略", "tier_multiplier/sku_override", "sku_override", "人工日均销目标值设定粒度"],
        ["P12", "daily_sale_tier_min_threshold", "日均销货盘最低阈值", "按S/A/B/C/N正整数", '{"S":0,"A":0,"B":5,"C":10,"N":1}', "低于阈值标记需人工介入"],
        ["P13", "promo_trigger_threshold", "促销触发阈值", "按S/A/B/C/N整数(可售天数超过此触发促销)", '{"S":45,"A":55,"B":65,"C":75,"N":90}', "场景06四区段预警-关注/预警边界"],
        ["P14", "reasonable_turnover_days", "合理周转天数", "按S/A/B/C/N整数", '{"S":30,"A":40,"B":50,"C":60,"N":75}', "场景06四区段预警-正常/关注边界"],
    ]

    wb = Workbook()
    ws = wb.active; ws.title = "参数配置"
    write_sheet(ws, headers, explains, params)
    return make_xlsx_response(wb, "参数配置.xlsx")


@app.get("/template/msu-map")
def template_msu_map():
    """SKU-最小销售单元-仓库映射模板"""
    sql = """select distinct concat(shop,site) as msu_id
from v_dwd_com_sell_1d
where pt2='亚马逊'
order by msu_id"""
    data = query(sql)

    headers = ["SKU编码", "最小销售单元ID", "供给仓库ID", "updated_at"]
    explains = ["内部统一SKU", "如CI-eu-DE德国（concat(shop,site)）", "海外仓库编码（cosboard可查但无现成映射表）", "数据更新时间"]
    gap_cols = {0, 2}  # SKU需关联，仓库需手动填

    rows = []
    for r in data:
        rows.append(["", r.get('msu_id',''), "", datetime.now().strftime("%Y-%m-%d")])

    wb = Workbook()
    ws = wb.active; ws.title = "MSU映射"
    write_sheet(ws, headers, explains, rows, gap_cols)
    return make_xlsx_response(wb, "SKU-最小销售单元-仓库映射.xlsx")


@app.get("/template/inventory")
def template_inventory():
    """库存快照 — 2 Sheet（国内 + 海外），对齐飞书模板"""
    # 国内库存
    inv = query("""select sku_id as sku, sum(invqty) as domestic_stock
from v_jstskuinv group by sku_id""")
    # 采购在途（归一化 SKU）
    pur = query("""select coalesce(b.sku_id, t.SkuCode) as sku, sum(t.onroadqty) as in_transit
from v_lmpurchase t
left join v_cdm_skubom b on t.SkuCode = b.psku
where t.onroadqty<>0 group by coalesce(b.sku_id, t.SkuCode)""")
    pur_map = {r['sku']: r['in_transit'] for r in pur}

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet A: 国内库存快照
    ws_dom = wb.create_sheet("国内库存快照")
    dom_headers = ["SKU编码", "国内库存", "采购在途", "快照时间"]
    dom_explains = ["唯一标识", "国内仓库存", "已下单未发货", "快照时间戳"]
    dom_rows = [[r['sku'], r.get('domestic_stock',0),
                 pur_map.get(r['sku'], 0), now] for r in inv]
    write_sheet(ws_dom, dom_headers, dom_explains, dom_rows)

    # Sheet B: 海外库存快照（从 dwd_sh_amz_fba_stock 读取）
    ws_ovs = wb.create_sheet("海外库存快照")
    ovs_headers = ["SKU编码", "店铺", "仓库", "可售库存", "在途库存", "快照时间"]
    ovs_explains = ["唯一标识", "仓库所属店铺（从仓库名解析）",
                    "海外仓名称", "当前可售件数", "在途件数",
                    "快照时间戳"]
    ovs_rows = []
    try:
        ovs_data = query("""
            SELECT b.sku_id AS sku_code, a.warehouse_name,
                   CAST(a.available AS DECIMAL(20,2)) * CAST(b.rm_qty AS DECIMAL(20,2)) AS available,
                   CAST(a.in_transit AS DECIMAL(20,2)) * CAST(b.rm_qty AS DECIMAL(20,2)) AS in_transit,
                   a.update_date
            FROM dwd_sh_amz_fba_stock a
            JOIN v_cdm_skubom b ON a.goods_sku = b.psku
            WHERE a.update_date = (SELECT MAX(update_date) FROM dwd_sh_amz_fba_stock)
            ORDER BY a.goods_sku
        """)
        for r in ovs_data:
            wh = r.get('warehouse_name','') or ''
            shop = wh.split('-')[0] if '-' in wh else wh
            ovs_rows.append([
                sanitize_cell(r.get('sku_code','')),
                sanitize_cell(shop), sanitize_cell(wh),
                int(r.get('available',0)), int(r.get('in_transit',0)),
                str(r.get('update_date',''))[:10],
            ])
    except Exception as e:
        print(f"  ⚠️ 海外库存查询失败: {e}")
    write_sheet(ws_ovs, ovs_headers, ovs_explains, ovs_rows, set())

    return make_xlsx_response(wb, "库存快照.xlsx")


@app.get("/template/daily-data")
def template_daily_data():
    """日增量数据：1 个 Excel，5 个 Sheet — 全部取上一个自然日"""
    return make_xlsx_response(build_daily_workbook(), "日增量数据.xlsx")


def build_daily_workbook() -> Workbook:
    """构建日增量数据多 Sheet 工作簿 — 全部取上一个自然日"""
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: 日销量
    ws1 = wb.create_sheet("日销量")
    sales = query("""select sku_id as sku, saledate, concat(shop,site) as msu_id, sum(qty) as qty
from v_dwd_com_sell_1d where pt2='亚马逊' and saledate = curdate() - interval 1 day
group by sku, saledate, msu_id order by saledate desc""")
    write_sheet(ws1, ["SKU编码", "销售日期", "日销量", "最小销售单元"],
                ["归一后的内部统一SKU", "YYYY-MM-DD", "当日销售件数", "最小销售单元标识"],
                [[r['sku'], str(r.get('saledate',''))[:10], r.get('qty',0), r.get('msu_id','')] for r in sales])

    # Sheet 2: 采购明细
    ws2 = wb.create_sheet("采购明细")
    po = query("""select t.`code` as po, coalesce(b.sku_id, t.skucode) as sku,
       t.OrderDate, t.qty, t.DeliverDate
from v_lmmoontime t
left join v_cdm_skubom b on t.skucode = b.psku
where t.OrderDate = curdate() - interval 1 day order by t.OrderDate desc""")
    write_sheet(ws2, ["采购单号", "SKU编码", "下单日期", "采购数量", "预计到货日期"],
                ["采购订单编号", "归一后的内部统一SKU", "YYYY-MM-DD", "下单件数", "预计到仓日期"],
                [[r['po'], r['sku'], str(r.get('OrderDate',''))[:10], r.get('qty',''),
                  str(r.get('DeliverDate',''))[:10]] for r in po])

    # Sheet 3: 采购收货明细
    ws3 = wb.create_sheet("采购收货明细")
    recv = query("""select t.POPlusCode, coalesce(b.sku_id, t.SkuCode) as sku,
       t.BillDate, t.qty, t.WarehouseName
from v_lmwhin t
left join v_cdm_skubom b on t.SkuCode = b.psku
where date(t.BillDate) = curdate() - interval 1 day order by t.BillDate desc""")
    write_sheet(ws3, ["采购单号", "SKU编码", "收货日期", "收货数量", "收货仓库"],
                ["关联采购单号", "归一后的内部统一SKU", "YYYY-MM-DD", "实际收到件数", "入仓仓库"],
                [[r['POPlusCode'], r['sku'], str(r.get('BillDate',''))[:19], r.get('qty',''),
                  r.get('WarehouseName','')] for r in recv])

    # Sheet 4: 海外发货明细（从 dwd_sh_amz_fba_shipment_sku_info 读取）
    ws4 = wb.create_sheet("海外发货明细")
    ship_rows = []
    try:
        ship_data = query("""
            SELECT b.sku_id AS sku_code,
                   CAST(a.shipping_quantity AS DECIMAL(20,2)) * CAST(b.rm_qty AS DECIMAL(20,2)) AS shipping_qty,
                   a.real_ship_date, a.expect_arrival_date,
                   a.shop_name, a.ship_sn, a.fulfillment_center
            FROM dwd_sh_amz_fba_shipment_sku_info a
            JOIN v_cdm_skubom b ON a.commodity_sku = b.psku
            WHERE DATE(a.real_ship_date) = CURDATE() - INTERVAL 1 DAY
              AND a.fulfillment_center IS NOT NULL AND a.fulfillment_center != ''
            ORDER BY a.real_ship_date DESC
        """)
        for r in ship_data:
            ship_rows.append([
                r.get('ship_sn',''),
                r.get('sku_code',''),
                str(r.get('real_ship_date',''))[:10],
                int(r.get('shipping_qty',0)),
                r.get('fulfillment_center',''),
                str(r.get('expect_arrival_date',''))[:10] if r.get('expect_arrival_date') else '',
                r.get('shop_name','')
            ])
    except Exception as e:
        print(f"  ⚠️ 海外发货查询失败: {e}")
    write_sheet(ws4,
                ["物流单号", "SKU编码", "发货日期", "发货数量", "收获仓库",
                 "预计到达日期", "店铺"],
                ["发货流水号", "商品SKU（cosboard原始编码）",
                 "实际发货日期", "发货件数", "FBA收货仓库代码",
                 "预计到货日期（用于在途计算）", "店铺名称"],
                ship_rows, set())

    # Sheet 5: 海外收货明细（仍为空 — cosboard 无实际收货日期字段）
    ws5 = wb.create_sheet("海外收货明细")
    write_sheet(ws5, ["物流单号", "SKU编码", "收货日期", "收货数量", "收货仓库"],
                ["发货流水号", "商品SKU（cosboard原始编码）",
                 "【cosboard无实际收货日期】仅有 expect_arrival_date（预计到达）",
                 "【cosboard无收货数量】发货单有 shipping_quantity，但无实际回传收货量",
                 "FBA收货仓库代码"],
                [], {2, 3})

    return wb


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8765)
