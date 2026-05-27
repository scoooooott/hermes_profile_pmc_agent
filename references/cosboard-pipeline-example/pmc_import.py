#!/usr/bin/env python3
"""PMC 数据管线 — Excel → DuckDB ODS 导入
三类文件 → 三目录 → 9 张 ODS 表
"""

import os, sys, re
from datetime import datetime

# ── 配置 ──────────────────────────────────────────
PMC_DIR = os.path.expanduser("~/pmc-data")
DB_PATH = os.path.expanduser("~/pmc-data/pmc_ods.duckdb")

# ── 工具函数 ──────────────────────────────────────
def sanitize(val):
    """清洗控制字符"""
    if isinstance(val, str):
        return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', val)
    return val

def read_excel(path):
    """读 Excel，跳过 Row1(表头)/Row2(解释)，返回 (headers, rows)"""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    return wb  # 返回整个 workbook，让调用方自行拿 Sheet

def import_sheet(con, wb, sheet_name, table_name, mode, upsert_keys=None):
    """导入单 Sheet → DuckDB 表
    
    mode: 'overwrite' = TRUNCATE + INSERT
          'upsert'    = INSERT ... ON CONFLICT DO NOTHING (去重追加)
    """
    if sheet_name not in wb.sheetnames:
        print(f"  ⚠️  Sheet '{sheet_name}' 不存在，跳过")
        return 0
    
    ws = wb[sheet_name]
    headers = [sanitize(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    
    # 读取 Row3+ 数据行
    rows = []
    for r in range(3, ws.max_row + 1):
        row = [sanitize(ws.cell(r, c).value) for c in range(1, len(headers) + 1)]
        # 跳过全空行
        if all(v is None or v == '' for v in row):
            continue
        rows.append(row)
    
    if not rows:
        print(f"  ⚠️  Sheet '{sheet_name}' 无数据，跳过")
        return 0
    
    # 生成列名（中文表头 → 英文列名）
    col_map = {
        "SKU编码": "sku_code", "SPU编码": "spu_code", "品名": "product_name",
        "类目": "category", "品牌": "brand", "上架日期": "launch_date",
        "上架状态": "status", "生命周期": "lifecycle", "货盘等级": "tier",
        "生产周期天数": "production_cycle_days",
        "人工日均销目标值": "manual_daily_sale_target",
        "MOQ(最小起订量)": "moq", "Lead Time(采购交期,天)": "lead_time",
        "updated_at": "updated_at",
        "销售日期": "sale_date", "日销量": "daily_qty",
        "最小销售单元": "msu_id", "最小销售单元ID": "msu_id",
        "采购单号": "po_number", "下单日期": "order_date",
        "采购数量": "order_qty", "预计到货日期": "eta",
        "收货日期": "receipt_date", "收货数量": "receipt_qty",
        "收货仓库": "warehouse_id", "供给仓库ID": "warehouse_id",
        "物流单号": "tracking_number", "发货日期": "ship_date",
        "发货数量": "ship_qty", "收获仓库": "dest_warehouse",
        "预计到达日期": "expect_arrival", "店铺": "shop",
        "可售库存": "inv_available", "锁定库存": "inv_locked",
        "在途库存": "inv_onway", "在途库存(采购在途)": "inv_onway",
        "国内库存": "inv_domestic", "采购在途": "inv_purchase_onway",
        "采购在单": "inv_inorder", "快照时间": "snapshot_time",
        "仓库列表": "warehouses", "店铺": "shop",
        "仓库": "warehouse_code",
        "参数编号": "param_no", "参数ID": "param_id", "参数名称": "param_name",
        "数据类型/取值范围": "param_type", "默认值": "param_default", "备注": "param_note",
    }
    
    cols = [col_map.get(h, f'col_{i}') for i, h in enumerate(headers)]
    
    # 替换 header 中的无效字符
    cols = [re.sub(r'[^a-z0-9_]', '_', c.lower()) for c in cols]
    
    # 临时表
    tmp = f"_tmp_{table_name}"
    
    if mode == 'overwrite':
        con.execute(f"DROP TABLE IF EXISTS {table_name}")
        # 用临时表 CREATE + INSERT 再 RENAME 避免中途崩溃丢表
        con.execute(f"DROP TABLE IF EXISTS {tmp}")
    
    # CREATE TABLE AS SELECT
    placeholders = ', '.join(['?' for _ in cols])
    col_defs = ', '.join(f'"{c}" VARCHAR' for c in cols)
    
    if mode == 'overwrite':
        con.execute(f'CREATE TABLE "{tmp}" ({col_defs})')
        con.executemany(f'INSERT INTO "{tmp}" VALUES ({placeholders})', rows)
        con.execute(f'DROP TABLE IF EXISTS "{table_name}"')
        con.execute(f'ALTER TABLE "{tmp}" RENAME TO "{table_name}"')
    else:
        # 先确保表存在
        existing = con.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name='{table_name}'").fetchone()
        if not existing:
            con.execute(f'CREATE TABLE "{table_name}" ({col_defs})')
        if upsert_keys:
            key_cols = ', '.join(f'"{col_map.get(k, k)}"' for k in upsert_keys)
            # DuckDB INSERT OR IGNORE 需要 UNIQUE 约束；改用 ON CONFLICT DO NOTHING
            try:
                con.executemany(f'INSERT INTO "{table_name}" VALUES ({placeholders}) ON CONFLICT ({key_cols}) DO NOTHING', rows)
            except Exception:
                # fallback: 先建唯一索引再 INSERT OR IGNORE
                con.execute(f'CREATE UNIQUE INDEX IF NOT EXISTS idx_{table_name}_upsert ON "{table_name}" ({key_cols})')
                con.executemany(f'INSERT OR IGNORE INTO "{table_name}" VALUES ({placeholders})', rows)
        else:
            con.executemany(f'INSERT INTO "{table_name}" VALUES ({placeholders})', rows)
    
    return len(rows)


# ── 主流程 ────────────────────────────────────────
def main():
    import duckdb
    con = duckdb.connect(DB_PATH)
    
    print(f"PMC DuckDB import @ {datetime.now().isoformat()}")
    print(f"DB: {DB_PATH}\n")
    
    results = {}
    
    # ═══ static/ 全量覆盖 ═══
    static_dir = f"{PMC_DIR}/static"
    static_files = {
        "商品主数据.xlsx": ("商品主数据", "ods_skus", "overwrite"),
        "参数配置.xlsx": ("参数配置", "ods_params", "overwrite"),
        "SKU-最小销售单元-仓库映射.xlsx": ("MSU映射", "ods_wmap", "overwrite"),
    }
    
    for fname, (sheet, table, mode) in static_files.items():
        path = f"{static_dir}/{fname}"
        if not os.path.exists(path):
            results[f"static/{fname}"] = "❌ 文件不存在"
            continue
        try:
            wb = read_excel(path)
            n = import_sheet(con, wb, sheet, table, mode)
            results[f"static/{fname}"] = f"✅ {n:,} rows → {table}"
        except Exception as e:
            results[f"static/{fname}"] = f"❌ {e}"
    
    # ═══ snapshot/ 全量覆盖 ═══
    snap_path = f"{PMC_DIR}/snapshot/库存快照.xlsx"
    if os.path.exists(snap_path):
        try:
            wb = read_excel(snap_path)
            n1 = import_sheet(con, wb, "国内库存快照", "ods_inventory_domestic", "overwrite")
            n2 = import_sheet(con, wb, "海外库存快照", "ods_inventory_overseas", "overwrite")
            results["snapshot/库存快照.xlsx"] = f"✅ 国内={n1:,} 海外={n2:,} rows"
        except Exception as e:
            results["snapshot/库存快照.xlsx"] = f"❌ {e}"
    else:
        results["snapshot/库存快照.xlsx"] = "❌ 文件不存在"
    
    # ═══ incremental/ 增量追加 ═══
    inc_path = f"{PMC_DIR}/incremental/日增量数据.xlsx"
    inc_sheets = {
        "日销量": ("ods_sales", "upsert", ("销售日期", "SKU编码", "最小销售单元")),
        "采购明细": ("ods_po", "upsert", ("下单日期", "采购单号", "SKU编码")),
        "采购收货明细": ("ods_po_recv", "upsert", ("收货日期", "采购单号", "SKU编码")),
        "海外发货明细": ("ods_ship", "upsert", ("发货日期", "物流单号", "SKU编码")),
        "海外收货明细": ("ods_ship_recv", "upsert", ("收货日期", "物流单号", "SKU编码")),
    }
    
    if os.path.exists(inc_path):
        try:
            wb = read_excel(inc_path)
            for sheet, (table, mode, keys) in inc_sheets.items():
                n = import_sheet(con, wb, sheet, table, mode, keys)
                results[f"inc/{sheet}"] = f"✅ {n:,} rows → {table}"
        except Exception as e:
            results["inc/日增量数据.xlsx"] = f"❌ {e}"
    else:
        results["inc/日增量数据.xlsx"] = "❌ 文件不存在"
    
    # ── 打印结果 ──
    print("=" * 60)
    for k, v in results.items():
        print(f"  {k}: {v}")
    
    # ── 表统计 ──
    print("\n─ 各表行数 ─")
    for row in con.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name").fetchall():
        t = row[0]
        cnt = con.execute(f'SELECT COUNT(*) FROM "{t}"').fetchone()[0]
        print(f"  {t}: {cnt:,} rows")
    
    con.close()
    print("\n✅ Import complete.")

if __name__ == "__main__":
    main()
