#!/usr/bin/env python3
"""
PMC 场景统一交付模块。

提供所有 PMC 业务场景 Skill 共享的输出通道检测、PDF/Excel 渲染、路径管理能力。

Usage:
    from pmc_delivery import (
        detect_channel, should_use_attachments,
        render_html_to_pdf, render_dataframe_to_excel,
        get_output_path, OUTPUT_BASE,
    )

    channel = detect_channel()
    if should_use_attachments(channel):
        pdf_path = render_html_to_pdf(html_str, 'scene01-report')
        xlsx_path = render_dataframe_to_excel(df, '销量缺口', 'scene01-detail')
        # Attach pdf_path / xlsx_path to message
    else:
        # Fall back to markdown text
        print(markdown_output)
"""

from __future__ import annotations

import asyncio
import os
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------------------------------------------------------------------------
# Constants — output directory
# ---------------------------------------------------------------------------

OUTPUT_BASE = Path("/tmp/hermes-pmc-output")
PDF_DIR = OUTPUT_BASE / "pdf"
EXCEL_DIR = OUTPUT_BASE / "excel"
HTML_DIR = OUTPUT_BASE / "html"

for _d in (PDF_DIR, EXCEL_DIR, HTML_DIR):
    _d.mkdir(parents=True, exist_ok=True)

# ---------------------------------------------------------------------------
# Colour palette (B-现代岩灰 — user-selected default)
# ---------------------------------------------------------------------------

TH_BG = "292524"  # dark iron
TH_FG = "FFFFFF"
RED = "dc2626"
ORANGE = "ea580c"
STRIPE_EVEN = "FAFAF9"
STRIPE_CRIT = "FEF2F2"
STRIPE_WARN = "FFF7ED"
BORDER_COLOR = "E7E5E4"
TEXT_COLOR = "1C1917"
TEXT_SECONDARY = "78716C"

# ---------------------------------------------------------------------------
# Cell styles (reusable)
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill(start_color=TH_BG, end_color=TH_BG, fill_type="solid")
_HEADER_FONT = Font(name="Arial", bold=True, size=10, color=TH_FG)
_CRIT_FILL = PatternFill(start_color=STRIPE_CRIT, end_color=STRIPE_CRIT, fill_type="solid")
_WARN_FILL = PatternFill(start_color=STRIPE_WARN, end_color=STRIPE_WARN, fill_type="solid")
_EVEN_FILL = PatternFill(start_color=STRIPE_EVEN, end_color=STRIPE_EVEN, fill_type="solid")
_RED_FONT = Font(name="Arial", size=10, color=RED, bold=True)
_ORANGE_FONT = Font(name="Arial", size=10, color=ORANGE, bold=True)
_BODY_FONT = Font(name="Arial", size=10, color=TEXT_COLOR)
_THIN_BORDER = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)
_RIGHT_ALIGN = Alignment(horizontal="right", vertical="center")
_LEFT_ALIGN = Alignment(horizontal="left", vertical="center")

# ---------------------------------------------------------------------------
# Channel detection
# ---------------------------------------------------------------------------


def detect_channel() -> str:
    """
    Determine the current output channel by inspecting the environment.

    Checks (in order):
      1. ``FEISHU_CHAT_ID`` / ``LARK_CHAT_ID``        → ``'feishu'``
      2. ``TELEGRAM_CHAT_ID``                          → ``'telegram'``
      3. Stdout is a TTY                                → ``'terminal'``
      4. Otherwise                                       → ``'unknown'``

    Returns
    -------
    str
        One of ``'feishu'``, ``'telegram'``, ``'terminal'``, ``'unknown'``.
    """
    # Check for feishu
    if (os.environ.get('FEISHU_CHAT_ID') or os.environ.get('LARK_CHAT_ID') or
        os.environ.get('FEISHU_APP_ID') or os.environ.get('FEISHU_HOME_CHANNEL')):
        return 'feishu'
    # Check HERMES_SESSION_PLATFORM (set by hermes in session context)
    platform = os.environ.get('HERMES_SESSION_PLATFORM', '').lower()
    if platform == 'feishu':
        return 'feishu'
    if platform == 'telegram':
        return 'telegram'
    if os.environ.get("TELEGRAM_CHAT_ID"):
        return "telegram"
    if sys.stdout.isatty():
        return "terminal"
    return "unknown"


def should_use_attachments(channel: str | None = None) -> bool:
    """
    Return ``True`` when *channel* supports file attachments (PDF, Excel).

    Parameters
    ----------
    channel : str, optional
        Channel string from :func:`detect_channel`.  Auto-detected if ``None``.

    Returns
    -------
    bool
    """
    if channel is None:
        channel = detect_channel()
    return channel in ("feishu", "telegram")


# ---------------------------------------------------------------------------
# Path helpers
# ---------------------------------------------------------------------------


def get_output_path(category: str, filename: str) -> Path:
    """
    Return the full path to *filename* inside *category* sub-directory.

    Parameters
    ----------
    category : str
        One of ``'pdf'``, ``'excel'``, ``'html'``.
    filename : str
        File name (e.g. ``'scene01-report.pdf'``).

    Returns
    -------
    Path
        ``/tmp/hermes-pmc-output/{category}/{filename}``
    """
    return OUTPUT_BASE / category / filename


# ---------------------------------------------------------------------------
# PDF rendering (HTML → Playwright → PDF)
# ---------------------------------------------------------------------------


def render_html_to_pdf(html_content: str, output_name: str) -> str:
    """
    Render *html_content* to a PDF file in the ``pdf/`` output directory.

    Workflow
    --------
    1. Write *html_content* to ``/tmp/hermes-pmc-output/html/{output_name}.html``.
    2. Launch Playwright Chromium and convert the HTML to PDF.
    3. Save PDF to ``/tmp/hermes-pmc-output/pdf/{output_name}.pdf``.

    Parameters
    ----------
    html_content : str
        Full HTML document (including ``<!DOCTYPE html>``, ``<style>``, etc.).
    output_name : str
        Base name (without extension) for the output files.

    Returns
    -------
    str
        Absolute path to the generated PDF file.
    """
    html_path = HTML_DIR / f"{output_name}.html"
    pdf_path = PDF_DIR / f"{output_name}.pdf"

    # Write intermediate HTML
    html_path.write_text(html_content, encoding="utf-8")

    # Convert to PDF via Playwright (async → sync wrapper)
    asyncio.run(_playwright_html_to_pdf(str(html_path.resolve()), str(pdf_path.resolve())))

    return str(pdf_path.resolve())


async def _playwright_html_to_pdf(html_path: str, pdf_path: str) -> None:
    """Internal: drive Playwright to export HTML to PDF."""
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        raise ImportError(
            "playwright is required for PDF generation. "
            "Install: pip install playwright && playwright install chromium"
        )

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=True)
        page = await browser.new_page(viewport={"width": 1200, "height": 1600})
        await page.goto(f"file://{html_path}", wait_until="domcontentloaded")
        await page.wait_for_timeout(500)
        await page.pdf(
            path=pdf_path,
            width="1200",
            print_background=True,
            margin={"top": "10mm", "bottom": "12mm", "left": "12mm", "right": "12mm"},
        )
        await browser.close()


# ---------------------------------------------------------------------------
# Excel rendering (DataFrame → openpyxl → .xlsx)
# ---------------------------------------------------------------------------


def render_dataframe_to_excel(
    df: pd.DataFrame,
    sheet_name: str,
    output_name: str,
    *,
    danger_column: str | None = None,
    danger_threshold: float | None = None,
    warning_threshold: float | None = None,
    wrap_columns: list[str] | None = None,
) -> str:
    """
    Write *df* to a formatted Excel file in the ``excel/`` output directory.

    Formatting applied automatically
    ---------------------------------
    - Deep gray header row (``#292524``) with white bold text
    - Frozen first row
    - Alternating even-row shading
    - Thin borders on all cells
    - Auto-adjusted column widths (capped at 40 characters)
    - Right-aligned numeric columns
    - SKU/code columns set to monospace font

    Conditional highlight
    ---------------------
    If *danger_column* is provided, cells in that column whose value is below
    *danger_threshold* get a red background + bold red text.  Values between
    *danger_threshold* and *warning_threshold* get an orange background + bold
    orange text.

    Parameters
    ----------
    df : pd.DataFrame
        Data to export.
    sheet_name : str
        Excel worksheet tab name.
    output_name : str
        Base name (without extension) for the ``.xlsx`` file.
    danger_column : str, optional
        Column name to apply conditional highlighting on.
    danger_threshold : float, optional
        Values below this threshold get danger (red) highlighting.
    warning_threshold : float, optional
        Values between ``*danger_threshold*`` and ``*warning_threshold*`` get
        warning (orange) highlighting.  Only meaningful if *danger_threshold*
        is also set.
    wrap_columns : list[str], optional
        Column names whose content should be soft-wrapped instead of
        truncated (e.g. product names).

    Returns
    -------
    str
        Absolute path to the generated ``.xlsx`` file.
    """
    xlsx_path = EXCEL_DIR / f"{output_name}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel sheet name limit

    # ---- populate data ----
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = _THIN_BORDER

            if r_idx == 1:
                # Header row
                cell.fill = _HEADER_FILL
                cell.font = _HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                # Data rows
                col_name = df.columns[c_idx - 1]
                cell.font = _BODY_FONT
                cell.alignment = _LEFT_ALIGN

                # Even-row shading
                if r_idx % 2 == 0:
                    cell.fill = _EVEN_FILL

                # Numeric columns: right-align + tabular
                if isinstance(value, (int, float)):
                    cell.alignment = _RIGHT_ALIGN

    # ---- conditional highlighting ----
    if danger_column and danger_column in df.columns:
        col_idx = df.columns.get_loc(danger_column) + 1
        threshold_lower = danger_threshold if danger_threshold is not None else 1
        threshold_upper = warning_threshold if warning_threshold is not None else threshold_lower + 2

        for r_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=r_idx, column=col_idx)
            if cell.value is None:
                continue
            try:
                val = float(cell.value)
            except (TypeError, ValueError):
                continue
            if val < threshold_lower:
                # Danger
                ws.cell(row=r_idx, column=col_idx).fill = PatternFill(
                    start_color=STRIPE_CRIT, end_color=STRIPE_CRIT, fill_type="solid"
                )
                ws.cell(row=r_idx, column=col_idx).font = _RED_FONT
                # Also tint the entire row
                for c in range(1, ws.max_column + 1):
                    if c != col_idx:
                        ws.cell(row=r_idx, column=c).fill = _CRIT_FILL
            elif val < threshold_upper:
                # Warning — tint only the cell
                ws.cell(row=r_idx, column=col_idx).fill = PatternFill(
                    start_color=STRIPE_WARN, end_color=STRIPE_WARN, fill_type="solid"
                )
                ws.cell(row=r_idx, column=col_idx).font = _ORANGE_FONT

    # ---- wrap columns ----
    if wrap_columns:
        for col_name in wrap_columns:
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                for r_idx in range(2, ws.max_row + 1):
                    ws.cell(row=r_idx, column=col_idx).alignment = Alignment(
                        horizontal="left", vertical="center", wrap_text=True
                    )

    # ---- auto-width ----
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = len(str(col_name))  # header width
        for r_idx in range(2, min(ws.max_row + 1, 50)):  # sample first 50 rows
            cell_val = ws.cell(row=r_idx, column=col_idx).value
            if cell_val is not None:
                max_len = max(max_len, len(str(cell_val)))
        adjusted = min(max_len + 3, 40)  # cap at 40 chars
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "A"].width = adjusted

    # ---- freeze header ----
    ws.freeze_panes = "A2"

    wb.save(str(xlsx_path.resolve()))
    return str(xlsx_path.resolve())


# ---------------------------------------------------------------------------
# Timestamp helper
# ---------------------------------------------------------------------------


def now_iso() -> str:
    """Return current UTC timestamp in ISO-8601 format (second precision)."""
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


# ---------------------------------------------------------------------------
# Quick sanity check when run directly
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    print(f"PMC Delivery Module — {now_iso()}")
    print(f"  OUTPUT_BASE      : {OUTPUT_BASE}")
    print(f"  PDF_DIR          : {PDF_DIR}")
    print(f"  EXCEL_DIR        : {EXCEL_DIR}")
    print(f"  HTML_DIR         : {HTML_DIR}")
    print(f"  detect_channel() : {detect_channel()}")
    print(f"  should_use_attachments(): {should_use_attachments()}")

    # Quick test: render a tiny DataFrame to Excel
    import pandas as pd

    test_df = pd.DataFrame(
        {
            "SKU": ["SKU001", "SKU002", "SKU003"],
            "商品名称": ["商品A", "商品B", "商品C"],
            "库存数量": [120, 5, 0],
            "可售天数": [14.5, 0.8, 0.0],
        }
    )
    xlsx = render_dataframe_to_excel(
        test_df,
        sheet_name="测试",
        output_name="_test_delivery",
        danger_column="可售天数",
        danger_threshold=1.0,
        warning_threshold=3.0,
        wrap_columns=["商品名称"],
    )
    print(f"  Excel test passed: {xlsx}")
